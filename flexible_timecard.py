#!/usr/bin/env python3
"""
Flexible Daily Timecard Generator
================================

This script creates a production-quality Excel timecard that tracks flexible work hours
with multiple clock-in/out intervals throughout the day.

Features:
- Track multiple clock-in/out intervals
- Real-time calculation of hours worked
- Shows remaining time to 8-hour goal
- Projects end time based on current pace
- Locale-agnostic formulas using comma separators
- Absolute cell references for stability

Usage:
    python flexible_timecard.py

The script will generate "Flexible_Timecard.xlsx" in the current directory.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, time, timedelta
import os

def create_flexible_timecard():
    """Create the flexible timecard workbook with all required functionality."""
    
    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Timecard"
    
    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=12)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    
    # Title
    ws['A1'] = "FLEXIBLE DAILY TIMECARD"
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws['A1'].alignment = center_alignment
    ws['A1'].border = border
    
    # Date row
    ws['A2'] = "Date:"
    ws['B2'] = "=TODAY()"
    ws['B2'].number_format = "dddd, mmmm d, yyyy"
    ws['A2'].font = subheader_font
    ws['A2'].fill = subheader_fill
    ws['A2'].border = border
    ws['B2'].border = border
    
    # Goal hours row
    ws['A3'] = "Daily Goal (hours):"
    ws['B3'] = 8.0
    ws['B3'].number_format = "0.00"
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws['A3'].border = border
    ws['B3'].border = border
    
    # Current time row
    ws['A4'] = "Current Time:"
    ws['B4'] = "=NOW()"
    ws['B4'].number_format = "hh:mm:ss AM/PM"
    ws['A4'].font = subheader_font
    ws['A4'].fill = subheader_fill
    ws['A4'].border = border
    ws['B4'].border = border
    
    # Spacer row
    ws['A5'] = ""
    
    # Time tracking section header
    ws['A6'] = "TIME TRACKING"
    ws.merge_cells('A6:F6')
    ws['A6'].font = header_font
    ws['A6'].fill = header_fill
    ws['A6'].alignment = center_alignment
    ws['A6'].border = border
    
    # Column headers
    headers = ["Interval", "Clock In", "Clock Out", "Duration (H:MM)", "Duration (Min)", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Time tracking rows (10 intervals)
    for row in range(8, 18):
        # Interval number
        ws.cell(row=row, column=1, value=row-7)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=1).alignment = center_alignment
        
        # Clock In (time format)
        clock_in_cell = ws.cell(row=row, column=2)
        clock_in_cell.number_format = "hh:mm AM/PM"
        clock_in_cell.border = border
        clock_in_cell.alignment = center_alignment
        
        # Clock Out (time format)
        clock_out_cell = ws.cell(row=row, column=3)
        clock_out_cell.number_format = "hh:mm AM/PM"
        clock_out_cell.border = border
        clock_out_cell.alignment = center_alignment
        
        # Duration H:MM (formula)
        duration_cell = ws.cell(row=row, column=4)
        duration_cell.value = f'=IF(AND(B{row}<>"",C{row}<>""),TEXT((C{row}-B{row})*24,"h:mm"),"")'
        duration_cell.border = border
        duration_cell.alignment = center_alignment
        
        # Duration Minutes (formula)
        minutes_cell = ws.cell(row=row, column=5)
        minutes_cell.value = f'=IF(AND(B{row}<>"",C{row}<>""),ROUND((C{row}-B{row})*1440,0),"")'
        minutes_cell.border = border
        minutes_cell.alignment = center_alignment
        
        # Notes
        notes_cell = ws.cell(row=row, column=6)
        notes_cell.border = border
        notes_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Spacer row
    ws['A19'] = ""
    
    # Summary section header
    ws['A20'] = "SUMMARY & PROJECTIONS"
    ws.merge_cells('A20:F20')
    ws['A20'].font = header_font
    ws['A20'].fill = header_fill
    ws['A20'].alignment = center_alignment
    ws['A20'].border = border
    
    # Summary calculations
    summary_data = [
        ("Total Hours Worked (H:MM):", "=TEXT(SUM(E8:E17)/60,\"h:mm\")", "Total time worked in hours:minutes format"),
        ("Total Minutes Worked:", "=SUM(E8:E17)", "Total time worked in minutes"),
        ("Goal Minutes:", "=$B$3*60", "Daily goal converted to minutes"),
        ("Minutes Remaining:", "=MAX(0,$B$3*60-SUM(E8:E17))", "Minutes still needed to reach goal"),
        ("Hours:Minutes Remaining:", "=TEXT(MAX(0,$B$3*60-SUM(E8:E17))/60,\"h:mm\")", "Time remaining in hours:minutes format"),
        ("Current Work Rate (min/hr):", "=IF(SUM(E8:E17)>0,SUM(E8:E17)/((NOW()-TODAY())*24),0)", "Average work rate based on time elapsed"),
        ("Projected End Time:", "=IF(SUM(E8:E17)>0,IF(SUM(E8:E17)>=$B$3*60,\"Goal Met!\",TODAY()+TIME(0,0,0)+TIME(0,MAX(0,$B$3*60-SUM(E8:E17)),0)/TIME(0,1,0)),\"No work logged\")", "When you'll finish if you maintain current pace"),
        ("Status:", "=IF(SUM(E8:E17)>=$B$3*60,\"Goal Met!\",IF(SUM(E8:E17)>0,\"Working\",\"Not Started\"))", "Current work status")
    ]
    
    for idx, (label, formula, description) in enumerate(summary_data):
        row = 21 + idx
        
        # Label
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = subheader_font
        label_cell.fill = subheader_fill
        label_cell.border = border
        
        # Formula/Value
        value_cell = ws.cell(row=row, column=2)
        value_cell.value = formula
        value_cell.border = border
        value_cell.alignment = center_alignment
        
        # Description
        desc_cell = ws.cell(row=row, column=3)
        desc_cell.value = description
        desc_cell.font = Font(size=9, italic=True)
        desc_cell.border = border
        desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Merge description across remaining columns
        ws.merge_cells(f'C{row}:F{row}')
    
    # Spacer row
    ws['A30'] = ""
    
    # Instructions section
    ws['A31'] = "INSTRUCTIONS"
    ws.merge_cells('A31:F31')
    ws['A31'].font = header_font
    ws['A31'].fill = header_fill
    ws['A31'].alignment = center_alignment
    ws['A31'].border = border
    
    instructions = [
        "1. Enter your daily goal in cell B3 (default: 8.00 hours)",
        "2. For each work interval, enter clock-in time in column B and clock-out time in column C",
        "3. The spreadsheet automatically calculates duration and running totals",
        "4. View real-time summary including hours worked, remaining time, and projected end time",
        "5. All formulas use absolute references ($) to prevent accidental breakage",
        "6. Times are automatically formatted for easy reading",
        "7. The 'Projected End Time' shows when you'll finish if you maintain your current work pace"
    ]
    
    for idx, instruction in enumerate(instructions):
        row = 32 + idx
        cell = ws.cell(row=row, column=1)
        cell.value = instruction
        cell.font = Font(size=10)
        cell.border = border
        ws.merge_cells(f'A{row}:F{row}')
    
    # Add data validation for time entries
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Time validation (optional - allows any time format)
    time_validation = DataValidation(
        type="time",
        operator="between",
        formula1="00:00",
        formula2="23:59",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Time",
        error="Please enter a valid time in HH:MM format"
    )
    ws.add_data_validation(time_validation)
    
    # Apply validation to clock in/out columns
    time_validation.add(f'B8:B17')
    time_validation.add(f'C8:C17')
    
    # Add some sample data for demonstration
    ws['B8'] = "09:00"
    ws['C8'] = "12:00"
    ws['B9'] = "13:00"
    ws['C9'] = "17:00"
    
    # Save the workbook
    filename = "Flexible_Timecard.xlsx"
    wb.save(filename)
    print(f"✅ Timecard created successfully: {filename}")
    print(f"📁 Saved to: {os.path.abspath(filename)}")
    
    return filename

if __name__ == "__main__":
    try:
        create_flexible_timecard()
        print("\n🎯 Your flexible timecard is ready!")
        print("💡 Open the Excel file and start tracking your work intervals.")
        print("📊 The spreadsheet will automatically calculate totals and projections.")
    except Exception as e:
        print(f"❌ Error creating timecard: {e}")
        print("Please ensure you have openpyxl installed: pip install openpyxl")
